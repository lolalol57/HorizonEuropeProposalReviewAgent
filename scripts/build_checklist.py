#!/usr/bin/env python3
"""Build the proposal-preparation checklist workbook for a run.

Renders a self-check workbook (3 sheets: Excellence / Impact / Implementation,
235 rubric checkpoints) from the canonical checkpoint list in
``scripts/data/checklist.json``.

  python3 scripts/build_checklist.py <run-id>              # filled if fill file present, else blank
  python3 scripts/build_checklist.py <run-id> --out PATH   # write to an explicit path
  python3 scripts/build_checklist.py <run-id> --blank      # force the blank template
  python3 scripts/build_checklist.py <run-id> --no-brand   # force the neutral (unbranded) look
  python3 scripts/build_checklist.py <run-id> --branding P # use branding profile at path P

When ``internal/checklist-fill.json`` exists, each checkpoint's Durum / Kanıt /
Notlar columns are filled from it (agent-filled review). Otherwise the blank
self-check template is produced.

Branding: the workbook title, subtitle, output filename and (optional) logo come
from a branding profile. The default is a neutral, unbranded look. A branding
profile may be supplied via ``--branding PATH``, the ``HE_REVIEW_BRANDING``
environment variable, or a ``branding/branding.json`` file in the working
directory / repo root — the first that exists wins.
"""
import argparse
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import common  # noqa: E402

CANONICAL = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "checklist.json")
SHEETS = ("Excellence", "Impact", "Implementation")

# review status -> Durum dropdown option (must match the data-validation list exactly)
STATUS_TO_DURUM = {
    "adequate": "✅ Addressed",
    "partial": "⚠️ Partially Addressed",
    "missing": "❌ Not Addressed",
    "na": "➖ N/A",
}
DURUM_OPTIONS = ["✅ Addressed", "⚠️ Partially Addressed", "❌ Not Addressed", "➖ N/A"]

# ---- branding -------------------------------------------------------------
# Neutral, unbranded defaults. A branding profile (JSON with any of these keys)
# can override title / subtitle / logo / output_filename.
NEUTRAL_BRANDING = {
    "title": "Proposal Preparation Checklist — {criterion}",
    "subtitle": ("Horizon Europe self-assessment checklist · {n} checkpoints · "
                 "mark each item and add evidence / section reference · "
                 "✅ Addressed / ⚠️ Partially Addressed / ❌ Not Addressed / ➖ N/A"),
    "logo": None,
    "output_filename": "06_Proposal_Preparation_Checklist.xlsx",
}


def resolve_branding(branding_arg=None, no_brand=False):
    """Return an effective branding dict (neutral defaults + optional overrides).

    Resolution order: --branding PATH, then $HE_REVIEW_BRANDING, then
    ./branding/branding.json (cwd, then repo root). --no-brand forces neutral.
    A relative ``logo`` in a profile is resolved against the profile's directory.
    """
    branding = dict(NEUTRAL_BRANDING)
    if no_brand:
        return branding

    candidates = []
    if branding_arg:
        candidates.append(branding_arg)
    env = os.environ.get("HE_REVIEW_BRANDING")
    if env:
        candidates.append(env)
    candidates.append(os.path.join(os.getcwd(), "branding", "branding.json"))
    candidates.append(os.path.join(common.REPO_ROOT, "branding", "branding.json"))

    path = next((p for p in candidates if p and os.path.exists(p)), None)
    if not path:
        return branding

    with open(path, encoding="utf-8") as f:
        profile = json.load(f)
    for key in ("title", "subtitle", "output_filename"):
        if profile.get(key):
            branding[key] = profile[key]
    logo = profile.get("logo")
    if logo:
        if not os.path.isabs(logo):
            logo = os.path.join(os.path.dirname(os.path.abspath(path)), logo)
        branding["logo"] = logo if os.path.exists(logo) else None
    return branding


# ---- palette / dimensions -------------------------------------------------
ACCENT = "E72229"
HEADERS = ["#", "Checkpoint", "Açıklama (Description)", "Durum",
           "Kanıt / Bölüm ref.", "Notlar / Aksiyon", "Sorumlu"]
COL_WIDTHS = {"A": 5.0, "B": 32.0, "C": 66.0, "D": 20.0, "E": 18.0, "F": 24.0, "G": 13.0}
CF_FILLS = {  # Durum value -> conditional-formatting fill
    "✅ Addressed": "C6EFCE",
    "⚠️ Partially Addressed": "FFEB9C",
    "❌ Not Addressed": "FFC7CE",
    "➖ N/A": "E7E7E7",
}

_thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _letterhead(ws, criterion, n_checkpoints, branding):
    """Logo band (optional) + title + subtitle."""
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    logo = branding.get("logo")
    # Row 1 — logo band (tall when a logo is present, a thin accent rule otherwise)
    if logo:
        ws.row_dimensions[1].height = 74.0
        ws.add_image(XLImage(logo), "A1")
    else:
        # Clean unbranded top: a thin accent rule instead of an empty tall band.
        ws.row_dimensions[1].height = 6.0
        ws.merge_cells("A1:G1")
        ws["A1"].fill = PatternFill("solid", fgColor=ACCENT)

    # Row 2 — title
    ws.merge_cells("A2:G2")
    t = ws["A2"]
    t.value = branding["title"].format(criterion=criterion)
    t.font = Font(name="Calibri", size=16, bold=True, color="222222")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 26.0

    # Row 3 — spacer
    ws.row_dimensions[3].height = 5.0

    # Row 4 — subtitle
    ws.merge_cells("A4:G4")
    s = ws["A4"]
    s.value = branding["subtitle"].format(n=n_checkpoints)
    s.font = Font(name="Calibri", size=9, italic=True, color="666666")
    s.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[4].height = 28.0


def _summary_band(ws, first_cp, last_row):
    labels = ["Toplam madde", "✅ Addressed", "⚠️ Partial", "❌ Not addr.", "➖ N/A"]
    band_fill = PatternFill("solid", fgColor="F2F2F2")
    for i, lab in enumerate(labels):
        c = ws.cell(row=5, column=i + 1, value=lab)
        c.font = Font(name="Calibri", size=8.5, bold=True, color="777777")
        c.fill = band_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    b_rng = "B%d:B%d" % (first_cp, last_row)
    d_rng = "D%d:D%d" % (first_cp, last_row)
    formulas = [
        "=COUNTA(%s)" % b_rng,
        '=COUNTIF(%s,"✅ Addressed")' % d_rng,
        '=COUNTIF(%s,"⚠️ Partially Addressed")' % d_rng,
        '=COUNTIF(%s,"❌ Not Addressed")' % d_rng,
        '=COUNTIF(%s,"➖ N/A")' % d_rng,
    ]
    for i, f in enumerate(formulas):
        c = ws.cell(row=6, column=i + 1, value=f)
        c.font = Font(name="Calibri", size=12, bold=True, color=ACCENT)
        c.fill = band_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _header_row(ws):
    fill = PatternFill("solid", fgColor="3A3A3A")
    for i, h in enumerate(HEADERS):
        c = ws.cell(row=7, column=i + 1, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[7].height = 22.0


def _group_row(ws, r, text):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    fill = PatternFill("solid", fgColor="F7D3D0")
    for col in range(1, 8):
        c = ws.cell(row=r, column=col)
        c.fill = fill
        c.border = BORDER
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(name="Calibri", size=10.5, bold=True, color=ACCENT)
    c.alignment = Alignment(horizontal="left", vertical="center")


def _checkpoint_row(ws, r, item, fill):
    num_al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

    a = ws.cell(row=r, column=1, value=item.get("num"))
    a.font = Font(name="Calibri", size=11, color="777777")
    a.alignment = num_al

    b = ws.cell(row=r, column=2, value=item.get("checkpoint"))
    b.font = Font(name="Calibri", size=10, color="222222")
    b.alignment = left_top

    c = ws.cell(row=r, column=3, value=item.get("description"))
    c.font = Font(name="Calibri", size=9.5, color="555555")
    c.alignment = left_top

    # D Durum / E Kanıt / F Notlar / G Sorumlu — filled if we have a mapping
    durum = citation = note = None
    if fill:
        status = fill.get("status")
        durum = STATUS_TO_DURUM.get(status) if status else None
        citation = fill.get("citation")
        note = fill.get("note")

    d = ws.cell(row=r, column=4, value=durum)
    d.font = Font(name="Calibri", size=11, color="222222")
    d.alignment = num_al

    e = ws.cell(row=r, column=5, value=citation)
    e.font = Font(name="Calibri", size=9, color="333333")
    e.alignment = left_top

    f = ws.cell(row=r, column=6, value=note)
    f.font = Font(name="Calibri", size=9, color="333333")
    f.alignment = left_top

    g = ws.cell(row=r, column=7)
    g.font = Font(name="Calibri", size=9, color="333333")
    g.alignment = num_al

    for col in range(1, 8):
        ws.cell(row=r, column=col).border = BORDER
    ws.row_dimensions[r].height = 48.0


def _build_sheet(ws, criterion, checkpoints, fill_map, branding):
    n = len(checkpoints)
    _letterhead(ws, criterion, n, branding)
    _header_row(ws)

    r = 8
    first_cp = None
    last_group = None
    for item in checkpoints:
        grp = item.get("group")
        if grp != last_group:
            _group_row(ws, r, grp)
            last_group = grp
            r += 1
        fill = (fill_map or {}).get(item.get("id"))
        _checkpoint_row(ws, r, item, fill)
        if first_cp is None:
            first_cp = r
        r += 1
    last_row = r - 1

    _summary_band(ws, first_cp, last_row)
    ws.freeze_panes = "A8"
    ws.auto_filter.ref = "A7:G%d" % last_row

    # Durum dropdown + conditional formatting on the checkpoint rows only
    d_range = "D%d:D%d" % (first_cp, last_row)
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(DURUM_OPTIONS), allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(d_range)
    for value, hexfill in CF_FILLS.items():
        ws.conditional_formatting.add(
            d_range,
            CellIsRule(operator="equal", formula=['"%s"' % value],
                       fill=PatternFill("solid", fgColor=hexfill)),
        )

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9  # A4


def build(checkpoints_by_sheet, fill_by_sheet, out_path, branding):
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in SHEETS:
        ws = wb.create_sheet(title=sheet)
        _build_sheet(ws, sheet, checkpoints_by_sheet[sheet],
                     (fill_by_sheet or {}).get(sheet), branding)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path


def _index_fill(fill_data):
    """Return {sheet: {checkpoint_id: {status,citation,note}}} from the fill JSON."""
    if not fill_data:
        return None
    out = {}
    for sheet in SHEETS:
        rows = fill_data.get(sheet) or []
        out[sheet] = {row["id"]: row for row in rows if row.get("id")}
    return out


def main(argv=None):
    ap = argparse.ArgumentParser(description="Build the proposal-preparation checklist workbook.")
    ap.add_argument("run_id")
    ap.add_argument("--out", help="output .xlsx path (default: OUTPUT/<branding output_filename>)")
    ap.add_argument("--blank", action="store_true", help="force the blank template (ignore any fill file)")
    ap.add_argument("--no-brand", action="store_true", help="force the neutral (unbranded) look")
    ap.add_argument("--branding", help="path to a branding profile JSON (overrides auto-detection)")
    args = ap.parse_args(argv)

    with open(CANONICAL, encoding="utf-8") as f:
        checkpoints_by_sheet = json.load(f)

    branding = resolve_branding(branding_arg=args.branding, no_brand=args.no_brand)

    fill_by_sheet = None
    if not args.blank:
        fill_data = common.read_json(common.internal_path(args.run_id, "checklist-fill.json"), default=None)
        fill_by_sheet = _index_fill(fill_data)

    out_path = args.out or os.path.join(
        common.subdir(args.run_id, "OUTPUT"), branding["output_filename"])
    build(checkpoints_by_sheet, fill_by_sheet, out_path, branding)

    kind = "blank template" if fill_by_sheet is None else "filled checklist"
    brand = "neutral" if branding.get("logo") is None and branding["title"] == NEUTRAL_BRANDING["title"] else "branded"
    total = sum(len(v) for v in checkpoints_by_sheet.values())
    print("Wrote %s (%s, %s, %d checkpoints across %d sheets)"
          % (out_path, kind, brand, total, len(SHEETS)))


if __name__ == "__main__":
    main()
